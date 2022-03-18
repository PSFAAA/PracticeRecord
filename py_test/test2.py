import openpyxl


class DoExcel:
    def __init__(self, file_name, sheet_name):  # file_name为你文件的路径，sheet_name为excel中表单名称
        self.file_name = file_name
        self.sheet_name = sheet_name

    # 将表格中的第一行的值作为键，其他行作为值，已字典的形式存在列表中
    def get_cases(self):
        wb = openpyxl.load_workbook(self.file_name)
        sh = wb[self.sheet_name]
        cases = []
        # 获取最大行号
        rows = sh.max_row
        #  获取最大列号
        cols = sh.max_column
        for row in range(2, rows + 1):
            case = {}
            for col in range(1, cols + 1):
                case[sh.cell(1, col).value] = sh.cell(row, col).value
                cases.append(case)
        wb.save(self.file_name)
        wb.close()
        return cases

    # row是你要传的行号，actual是你要断言的值，result是你要写入的用例执行结果
    # def write_result(self, row, actual, result):
    #     wb = openpyxl.load_workbook(self.file_name)
    #     sh = wb[self.sheet_name]
    #     sh.cell(row=row, column=8).value = actual  # 在第几行几列写入你的断言值
    #     sh.cell(row=row, column=9).value = result  # 在第几行几列写入你的用例运行结果，也就是拿你的断言值去与运行后获取到的某个值进行对比
    #     wb.save(self.file_name)
    #     wb.close()


if __name__ == '__main__':
    excel = DoExcel(r"C:\Users\sh.pan.tim\Desktop\PPP\阿里巴巴2020年股票数据.xlsx", "股票数据")
    print(excel.get_cases())
    # file_name为你文件的路径，sheet_name为excel中表单名称
    # datas = excel.get_cases()
    # if student_name[0] == data["name"]:
    #     excel.write_result(data["id"] + 1, student_name[0], "pass")
    #     # id为什么要加1，是因为表格中第一行是title，要在第二行写数据的时候，其实他的id是1，不过你的id也可以直接从2开始就不用加1了
    # else:
    #     excel.write_result(data["id"] + 1, student_name[0], "fail")
